from __future__ import annotations

import json
import math
import re
import time
import textwrap
import zipfile
from collections import Counter, OrderedDict
from datetime import date, datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[1]
GENERATED_DATE = date(2026, 5, 19).isoformat()
SCHEMA = "pindou-color-palette"
FIXED_DATETIME = datetime(2026, 5, 19, 0, 0, 0)
FIXED_ZIP_DATE = (2026, 5, 19, 0, 0, 0)
FIXED_W3CDTF = "2026-05-19T00:00:00Z"
FIXED_PDF_TIME = time.strptime("20260519000000", "%Y%m%d%H%M%S")


MARKET_RESEARCH_SOURCES = [
    {
        "name": "中工网/北京青年报：年轻人涌进 DIY 拼豆店",
        "url": "https://www.workercn.cn/c/2026-02-25/8740893.shtml",
        "signal": "报道提到 COCO 手工店价格；MARD 店铺粉丝、复购、抢购；玩家称 MARD 色号体系逐渐成为默认参考标准。",
    },
    {
        "name": "淮南日报 PDF：拼豆热升温，手作带火体验经济",
        "url": "https://hnrb.huainannet.com/attachment/202604/08/40d9546e-d38a-43b4-b314-b4b300b22616.pdf",
        "signal": "报道 2025-2026 年拼豆消费热度，并引用 Mard 月均产量和供应链增长信息。",
    },
    {
        "name": "比特拼豆色卡",
        "url": "https://bitbead.pomodiary.com/zh/colors",
        "signal": "将 Perler、Hama、Artkal、MARD 称为四大主流品牌；称 MARD 221 是国内零售最常见版本。",
    },
    {
        "name": "拼豆工具站 Mard 标准色卡/售卖页",
        "url": "https://www.pindou.online/colors",
        "signal": "以 Mard 标准色卡作为站内核心色卡；售卖页说明豆子与 221 色卡完全对应。",
    },
    {
        "name": "PinDou 图纸生成器",
        "url": "https://pindou-e90.pages.dev/",
        "signal": "覆盖 Perler、Hama、Artkal、MARD、COCO、漫漫、盼盼、咪小窝等 8 大品牌。",
    },
    {
        "name": "BeadPattern 拼豆图纸生成器",
        "url": "https://beadpattern.net/zh/about",
        "signal": "支持 MARD、COCO、漫漫、盼盼、Artkal 等 5 大品牌色号系统。",
    },
    {
        "name": "爱拼豆 App Store",
        "url": "https://apps.apple.com/cn/app/%E7%88%B1%E6%8B%BC%E8%B1%86/id6756688364",
        "signal": "支持 20+ 种主流拼豆品牌色板，列出 Artkal、MARD、可可、慢慢家、盼盼等。",
    },
    {
        "name": "拼豆酱 Perler Chan",
        "url": "https://pindou.baby/",
        "signal": "面向中国拼豆玩家，内置 Mard、优肯、黄豆豆色卡；页面标注 Mard 为国内主流，黄豆豆新手友好。",
    },
    {
        "name": "公开源码 get-colors-from-beans",
        "url": "https://git.xiongxiao.me/abearxiong/get-colors-from-beans",
        "signal": "公开库覆盖 COCO、Mard、优肯、咪小窝、小舞、漫漫、盼盼、黄豆豆等色卡。",
    },
]

MARKET_ASSESSMENTS = {
    "coco-291": {
        "tier": "A-",
        "score": 3.9,
        "label": "常见性价比品牌",
        "summary": "COCO/可可在线下和电商语境中可见度高，常作为 MARD 体系之外的平价常用豆；图纸工具覆盖也较稳定。",
        "evidence": [
            "中工网报道引用电商平台数据提到“COCO手工店”10g 约 1000 粒拼豆售价。",
            "PinDou、BeadPattern、拼豆糕手等图纸工具均把 COCO 列入可选国内品牌。",
            "公开源码库提供 COCO 多规格色卡，说明其被工具生态持续收录。",
        ],
    },
    "mard-291-github": {
        "tier": "S",
        "score": 4.7,
        "label": "国内默认参考体系的完整版",
        "summary": "MARD 是国内玩家和手作小店最容易遇到的参考体系；291 色是 221 色基础上的扩展，全色制作用途强。",
        "evidence": [
            "中工网报道中玩家称 MARD 色号体系逐渐成为默认参考标准，许多图纸和教程会标注 MARD 色号。",
            "比特拼豆将 MARD 221/291 与 Artkal 等列入主流色卡，并称 MARD 221 是国内零售最常见版本。",
            "拼豆工具站、豆豆工坊、Alfonse 等都以 MARD 作为核心色卡或默认色卡。",
        ],
    },
    "mard-221-github": {
        "tier": "S",
        "score": 5.0,
        "label": "国内最主流/默认参考",
        "summary": "MARD 221 是国内零售和图纸工具最常见的标准版本之一，适合作为手工小店备货和图纸交流的基准色卡。",
        "evidence": [
            "比特拼豆页面直接称 MARD 221 是国内零售最常见的版本。",
            "中工网报道提到 MARD 店铺粉丝、复购和抢购，并引用玩家称其色号体系成默认参考。",
            "拼豆工具站售卖页说明豆子与 Mard 221 色卡完全对应，可直接按色号选购。",
        ],
    },
    "mard-221-alfonse-doudou": {
        "tier": "S",
        "score": 5.0,
        "label": "国内最主流/默认参考",
        "summary": "这是 MARD 221 的另一套公开工具站核对版；主流程度与 MARD 221 等同，建议作为人类对照优先版本之一。",
        "evidence": [
            "豆豆工坊页面写明 Mard 色卡是其拼豆工具默认色卡，共 221 个颜色。",
            "Alfonse 页面内嵌 MARD 官方拼豆色卡，用于 2.6mm MARD 拼豆像素制作。",
            "MARD 221 在多个图纸工具、售卖页和媒体报道中被作为标准体系提及。",
        ],
    },
    "panpan-289": {
        "tier": "B+",
        "score": 3.6,
        "label": "工具生态常见品牌",
        "summary": "盼盼在国内图纸工具和公开色卡中出现频率较高，但公开电商/媒体声量弱于 MARD、COCO、Artkal。",
        "evidence": [
            "PinDou 图纸生成器覆盖盼盼。",
            "BeadPattern 支持 MARD、COCO、漫漫、盼盼、Artkal 五大品牌色号系统。",
            "爱拼豆 App Store 描述列出盼盼拼豆等主流色板。",
        ],
    },
    "mixiaowo-290": {
        "tier": "B",
        "score": 3.4,
        "label": "常见但偏工具/玩家圈",
        "summary": "咪小窝在国内图纸工具中很常见，也有二手/材料包可见度；整体主流程度低于 MARD 和 COCO。",
        "evidence": [
            "PinDou 图纸生成器覆盖咪小窝。",
            "拼豆糕手 App 描述支持 MARD、COCO、漫漫、盼盼、咪小窝五大品牌。",
            "公开搜索可见咪小窝色号贴、套装与清仓信息。",
        ],
    },
    "xiaowu-291": {
        "tier": "B-",
        "score": 3.1,
        "label": "平价补充品牌",
        "summary": "小舞家在平价材料包和通用色卡贴中可见，公开源码也收录完整色卡；但主流工具直接支持和媒体声量少于 COCO/漫漫/盼盼/咪小窝。",
        "evidence": [
            "电商/代购搜索结果中常见 MARD、黄豆豆、DODO、小舞、COCO 同款材料包描述。",
            "公开源码库收录小舞 291 色完整色卡。",
            "选购文章把小舞家列为 1.3-1.4 元/1000 颗价位段的性价比选择。",
        ],
    },
    "manman-278": {
        "tier": "B+",
        "score": 3.7,
        "label": "老牌/图纸生态常见",
        "summary": "漫漫/慢慢在早期拼豆图纸和工具生态中存在感较高，常被图纸生成器列入国内主流品牌；电商声量中等。",
        "evidence": [
            "PinDou 和 BeadPattern 均支持漫漫/慢慢色号系统。",
            "拼豆糕手 App 描述支持 MARD、COCO、漫漫、盼盼、咪小窝五大品牌。",
            "入门/教程类页面称漫漫家是拼豆圈启蒙老店、图纸资源丰富。",
        ],
    },
    "huangdoudou-291": {
        "tier": "B",
        "score": 3.5,
        "label": "新手/平价常见品牌",
        "summary": "黄豆豆在材料包、通用色卡贴和新手工具中常见，适合作为平价/入门备选；图纸工具覆盖不如 MARD/COCO/漫漫/盼盼/咪小窝稳定。",
        "evidence": [
            "拼豆酱内置 Mard、优肯、黄豆豆色卡，并将黄豆豆标注为新手友好。",
            "多个商品标题把黄豆豆与 MARD、DODO、COCO、小舞同列为同款/通用体系。",
            "公开源码库收录黄豆豆 291 色完整色卡。",
        ],
    },
    "youken-public-174": {
        "tier": "B-",
        "score": 3.0,
        "label": "旧公开表/参考价值高于采购价值",
        "summary": "优肯/Artkal 本身较主流，但此 174 色公开库旧表不是当前最完整官方体系；适合交叉参考，不建议作为主备货口径。",
        "evidence": [
            "Artkal 在比特拼豆和多个工具中被列为主流或常见品牌。",
            "该表来自公开源码库的旧优肯数据，和官方 C197/M221 口径不完全一致。",
            "当前采购和对色更应看 Artkal 官方 C197、M221 或 C+M418。",
        ],
    },
    "youken-mard-221-public": {
        "tier": "A-",
        "score": 3.8,
        "label": "MARD 兼容补充体系",
        "summary": "优肯 MARD 同款 221 色承接 MARD 色号体系的需求，适合已经按 MARD 做图纸但想用优肯/Artkal 供应链的场景。",
        "evidence": [
            "抖音搜索结果可见“优肯拼豆推出 Mard 同款 221 个色号，加上之前 197 个色号共 418 个色号”的内容摘要。",
            "Artkal M 系列官方页写明 M-2.6mm 与 C 系列同材质同尺寸，并提供 221 色 RGB chart。",
            "公开源码库收录优肯Mard-221 色卡。",
        ],
    },
    "artkal-c-197-official": {
        "tier": "A",
        "score": 4.1,
        "label": "官方稳定/进阶常用",
        "summary": "Artkal/优肯是国内外都可见的成熟品牌；C 系列 197 色有官方色卡和 RGB PDF，适合对官方出处和颜色稳定性要求更高的玩家/小店。",
        "evidence": [
            "Artkal 官方 C-2.6mm 页面写明 197 colors，并提供 RGB color chart。",
            "比特拼豆将 Artkal 与 MARD 等列入主流色卡体系。",
            "拼豆酱和爱拼豆等工具把 Artkal/优肯纳入国内品牌色卡支持。",
        ],
    },
    "artkal-m-221-official": {
        "tier": "A-",
        "score": 3.9,
        "label": "官方 MARD 兼容新体系",
        "summary": "Artkal M 221 是官方 M 系列，对 MARD 色号需求有兼容意义；国内小店主流度目前低于 MARD 标准本体，但官方性和扩展价值较强。",
        "evidence": [
            "Artkal M-2.6mm 官方页写明 221 colors，并提供 RGB color chart。",
            "官方页说明 M 系列与 C 系列同材质同尺寸，可一起熔合。",
            "优肯/Artkal 在多个工具和教程中被列为主流或常见品牌。",
        ],
    },
    "artkal-c197-m221-418-official": {
        "tier": "A",
        "score": 4.2,
        "label": "进阶全量/官方体系",
        "summary": "优肯/Artkal 418 色是 C197 + M221 的官方合并口径，色域和官方出处强；对普通手工小店不是最便宜默认选择，但对进阶备货和精细图纸很有价值。",
        "evidence": [
            "Artkal 官方分别发布 C197 和 M221 RGB chart。",
            "比特拼豆称 Artkal 色号丰富，并把 Artkal/MARD 纳入主流色卡体系。",
            "拼豆酱面向中国玩家列出优肯 Artkal，并标注其色彩丰富。",
        ],
    },
}


HUE_GROUPS = [
    ("01-灰白黑", "灰白黑"),
    ("02-红粉", "红粉"),
    ("03-橙棕", "橙棕"),
    ("04-黄", "黄"),
    ("05-绿", "绿"),
    ("06-青蓝", "青蓝"),
    ("07-紫", "紫"),
    ("08-透明或未知", "透明或未知"),
]


def load_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for item in candidates:
        if Path(item).exists():
            return ImageFont.truetype(item, size=size)
    return ImageFont.load_default()


FONT_TITLE = load_font(34)
FONT_SUBTITLE = load_font(22)
FONT_BODY = load_font(18)
FONT_SMALL = load_font(14)
FONT_TINY = load_font(12)
FONT_TILE_CODE = load_font(20)
FONT_TILE_SMALL = load_font(12)


def coerce_int(value: Any) -> int | None:
    if value == "" or value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def clean_hex(value: Any) -> str | None:
    if not value:
        return None
    text = str(value).strip().upper()
    if re.fullmatch(r"#[0-9A-F]{6}", text):
        return text
    return None


def rgb_tuple(row: dict[str, Any]) -> tuple[int, int, int] | None:
    r = coerce_int(row.get("r"))
    g = coerce_int(row.get("g"))
    b = coerce_int(row.get("b"))
    if r is None or g is None or b is None:
        return None
    if not all(0 <= n <= 255 for n in (r, g, b)):
        return None
    return (r, g, b)


def hex_from_rgb(rgb: tuple[int, int, int] | None) -> str | None:
    if rgb is None:
        return None
    return "#{:02X}{:02X}{:02X}".format(*rgb)


def luminance(rgb: tuple[int, int, int] | None) -> float:
    if rgb is None:
        return 240
    r, g, b = rgb
    return (r * 299 + g * 587 + b * 114) / 1000


def text_color_for(rgb: tuple[int, int, int] | None) -> str:
    return "000000" if luminance(rgb) > 150 else "FFFFFF"


def excel_argb(rgb: tuple[int, int, int] | None) -> str:
    if rgb is None:
        return "FFD9D9D9"
    return "FF{:02X}{:02X}{:02X}".format(*rgb)


def code_prefix(code: str) -> str:
    match = re.match(r"^([A-Za-z]+)", code)
    return match.group(1).upper() if match else ""


def hue_bucket(row: dict[str, Any]) -> str:
    rgb = rgb_tuple(row)
    if rgb is None or row.get("transparency"):
        return "08-透明或未知"
    r, g, b = rgb
    if max(rgb) - min(rgb) < 28:
        return "01-灰白黑"

    import colorsys

    h, s, v = colorsys.rgb_to_hsv(r / 255, g / 255, b / 255)
    hue = h * 360
    if s < 0.12:
        return "01-灰白黑"
    if hue < 18 or hue >= 335:
        return "02-红粉"
    if hue < 48:
        return "03-橙棕"
    if hue < 75:
        return "04-黄"
    if hue < 165:
        return "05-绿"
    if hue < 255:
        return "06-青蓝"
    if hue < 335:
        return "07-紫"
    return "08-透明或未知"


def sheet_name(base: str, used: set[str], index: int) -> str:
    clean = re.sub(r"[\[\]\*:/\\?]", "-", base).strip() or "Sheet"
    clean = f"{index:02d}_{clean}"
    clean = clean[:31]
    candidate = clean
    suffix = 1
    while candidate in used:
        marker = f"_{suffix}"
        candidate = clean[: 31 - len(marker)] + marker
        suffix += 1
    used.add(candidate)
    return candidate


def rgb_text(row: dict[str, Any]) -> str:
    if row["rgb"] is None:
        return "RGB: N/A"
    return f"RGB: {row['r']}, {row['g']}, {row['b']}"


def rgb_from_color(row: dict[str, Any]) -> tuple[int, int, int] | None:
    rgb = row.get("rgb")
    if rgb is None:
        return None
    if isinstance(rgb, dict):
        values = [rgb.get("r"), rgb.get("g"), rgb.get("b")]
    else:
        values = list(rgb)
    if len(values) != 3:
        return None
    try:
        r, g, b = [int(v) for v in values]
    except (TypeError, ValueError):
        return None
    if not all(0 <= n <= 255 for n in (r, g, b)):
        return None
    return (r, g, b)


def normalize_xlsx_file(path: Path) -> None:
    tmp_path = path.with_suffix(".xlsx.tmp")
    ET.register_namespace("cp", "http://schemas.openxmlformats.org/package/2006/metadata/core-properties")
    ET.register_namespace("dc", "http://purl.org/dc/elements/1.1/")
    ET.register_namespace("dcterms", "http://purl.org/dc/terms/")
    ET.register_namespace("dcmitype", "http://purl.org/dc/dcmitype/")
    ET.register_namespace("xsi", "http://www.w3.org/2001/XMLSchema-instance")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        tmp_path,
        "w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=9,
    ) as target:
        for item in sorted(source.infolist(), key=lambda entry: entry.filename):
            data = source.read(item.filename)
            if item.filename == "docProps/core.xml":
                root = ET.fromstring(data)
                namespaces = {
                    "cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
                    "dcterms": "http://purl.org/dc/terms/",
                    "xsi": "http://www.w3.org/2001/XMLSchema-instance",
                }
                for tag in ("created", "modified"):
                    node = root.find(f"dcterms:{tag}", namespaces)
                    if node is not None:
                        node.text = FIXED_W3CDTF
                        node.set(f"{{{namespaces['xsi']}}}type", "dcterms:W3CDTF")
                data = ET.tostring(root, encoding="utf-8", xml_declaration=False)

            info = zipfile.ZipInfo(item.filename, date_time=FIXED_ZIP_DATE)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = item.external_attr
            target.writestr(info, data)
    tmp_path.replace(path)


def normalize_rows(raw_rows: list[dict[str, Any]], sources: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    source_map = {src.get("id"): src for src in (sources or [])}
    rows = []
    for idx, item in enumerate(raw_rows, start=1):
        rgb = rgb_from_color(item)
        source_ref = item.get("source") or ""
        source = source_map.get(source_ref, {})
        row = {
            "ordinal": idx,
            "code": str(item.get("code", "")).strip(),
            "hex": clean_hex(item.get("hex")) or hex_from_rgb(rgb),
            "rgb": None if rgb is None else {"r": rgb[0], "g": rgb[1], "b": rgb[2]},
            "r": None if rgb is None else rgb[0],
            "g": None if rgb is None else rgb[1],
            "b": None if rgb is None else rgb[2],
            "transparency": str(item.get("transparency", "") or ""),
            "source": source_ref,
            "source_quality": str(source.get("quality") or item.get("source_quality") or ""),
            "source_url": str(source.get("url") or item.get("source_url") or ""),
            "notes": str(item.get("notes") or source.get("notes") or ""),
            "group": str(item.get("group") or ""),
        }
        rows.append(row)
    return rows


def source_urls(rows: list[dict[str, Any]]) -> list[str]:
    urls = OrderedDict()
    for row in rows:
        url = str(row.get("source_url", "")).strip()
        if url:
            urls[url] = True
    return list(urls.keys())


def group_rows(rows: list[dict[str, Any]]) -> OrderedDict[str, list[dict[str, Any]]]:
    grouped: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
    if all(row.get("group") for row in rows):
        for row in rows:
            grouped.setdefault(str(row["group"]), []).append(row)
        return grouped

    prefixes = [code_prefix(str(row["code"])) for row in rows]
    non_empty_ratio = sum(1 for p in prefixes if p) / max(len(prefixes), 1)
    if non_empty_ratio >= 0.55:
        for row in rows:
            group = code_prefix(str(row["code"])) or hue_bucket(row)
            grouped.setdefault(group, []).append(row)
        return grouped

    hue_grouped = {key: [] for key, _ in HUE_GROUPS}
    for row in rows:
        hue_grouped[hue_bucket(row)].append(row)
    for key, label in HUE_GROUPS:
        if hue_grouped[key]:
            grouped[label] = hue_grouped[key]
    return grouped


def compact_sources(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: OrderedDict[tuple[str, str, str], str] = OrderedDict()
    result = []
    for row in rows:
        key = (row.get("source_url", ""), row.get("source_quality", ""), row.get("notes", ""))
        if key not in seen:
            source_id = f"src{len(seen) + 1}"
            seen[key] = source_id
            result.append({"id": source_id, "url": key[0], "quality": key[1], "notes": key[2]})
        row["source"] = seen[key]
    return result


def make_json(out_dir: Path, info: dict[str, Any], rows: list[dict[str, Any]], groups: OrderedDict[str, list[dict[str, Any]]]) -> None:
    market = info.get("market", {})
    sources = compact_sources(rows)
    colors = []
    for row in rows:
        rgb = None if row["rgb"] is None else [row["r"], row["g"], row["b"]]
        item = {
            "code": row["code"],
            "hex": row["hex"],
            "rgb": rgb,
            "group": row.get("group") or next((name for name, group in groups.items() if row in group), ""),
            "source": row.get("source") or None,
        }
        if row.get("transparency"):
            item["transparency"] = row["transparency"]
        if row.get("notes") and len(sources) > 1:
            item["notes"] = row["notes"]
        colors.append(item)
    payload = {
        "schema": SCHEMA,
        "id": info["id"],
        "title": info["title"],
        "description": info.get("description", info.get("long_title", info["title"])),
        "generated_at": GENERATED_DATE,
        "count": len(rows),
        "rows_with_rgb": sum(1 for row in rows if row["rgb"] is not None),
        "rows_without_rgb": sum(1 for row in rows if row["rgb"] is None),
        "market": market,
        "groups": {name: len(group) for name, group in groups.items()},
        "sources": sources,
        "colors": colors,
    }
    (out_dir / "colors.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def make_xlsx(out_dir: Path, info: dict[str, str], rows: list[dict[str, Any]], groups: OrderedDict[str, list[dict[str, Any]]]) -> None:
    market = info.get("market", {})
    wb = Workbook()
    wb.properties.creator = "HansBug"
    wb.properties.lastModifiedBy = "HansBug"
    wb.properties.created = FIXED_DATETIME
    wb.properties.modified = FIXED_DATETIME
    wb.remove(wb.active)
    info_ws = wb.create_sheet("信息")
    info_ws["A1"] = info["title"]
    info_ws["A1"].font = Font(bold=True, size=16)
    info_ws["A3"] = "生成日期"
    info_ws["B3"] = GENERATED_DATE
    info_ws["A4"] = "系列短名"
    info_ws["B4"] = info["title"]
    info_ws["A5"] = "完整标题"
    info_ws["B5"] = info.get("long_title", info["title"])
    info_ws["A6"] = "总颜色数"
    info_ws["B6"] = len(rows)
    info_ws["A7"] = "有 RGB"
    info_ws["B7"] = sum(1 for row in rows if row["rgb"] is not None)
    info_ws["A8"] = "无 RGB / 透明或未公开"
    info_ws["B8"] = sum(1 for row in rows if row["rgb"] is None)
    info_ws["A9"] = "来源"
    info_ws["A9"].font = Font(bold=True)
    info_ws["A10"] = "国内小店主流度"
    info_ws["B10"] = f"{market.get('tier', 'N/A')} / {market.get('score', 'N/A')} - {market.get('label', '')}"
    info_ws["A11"] = "评估摘要"
    info_ws["B11"] = market.get("summary", "")
    info_ws["B11"].alignment = Alignment(wrap_text=True, vertical="top")
    info_ws["A13"] = "来源"
    for row_idx, url in enumerate(source_urls(rows), start=14):
        info_ws.cell(row=row_idx, column=1, value=url)
    info_ws.column_dimensions["A"].width = 34
    info_ws.column_dimensions["B"].width = 70

    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    used_names = {"信息"}
    for idx, (group_name, group) in enumerate(groups.items(), start=1):
        ws = wb.create_sheet(sheet_name(group_name, used_names, idx))
        ws["A1"] = f"{info['title']} - {group_name}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"{len(group)} 色；每个色块含色号、HEX、RGB。"
        ws["A2"].font = Font(color="666666")
        ws.freeze_panes = "A4"

        tile_cols = 4
        tile_rows = 5
        cards_per_row = 4
        start_row = 4
        start_col = 1

        max_col = cards_per_row * tile_cols + (cards_per_row - 1)
        for col in range(1, max_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12

        for color_idx, row in enumerate(group):
            block_row = start_row + (color_idx // cards_per_row) * tile_rows
            block_col = start_col + (color_idx % cards_per_row) * (tile_cols + 1)
            rgb = rgb_tuple(row)
            fill = PatternFill("solid", fgColor=excel_argb(rgb))
            font_color = text_color_for(rgb)
            swatch_range = (
                block_row,
                block_col,
                block_row + 1,
                block_col + tile_cols - 1,
            )
            ws.merge_cells(
                start_row=swatch_range[0],
                start_column=swatch_range[1],
                end_row=swatch_range[2],
                end_column=swatch_range[3],
            )
            cell = ws.cell(row=block_row, column=block_col, value=row["code"])
            cell.fill = fill
            cell.font = Font(bold=True, color=font_color, size=14)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            for r in range(swatch_range[0], swatch_range[2] + 1):
                ws.row_dimensions[r].height = 24
                for c in range(swatch_range[1], swatch_range[3] + 1):
                    ws.cell(row=r, column=c).fill = fill
                    ws.cell(row=r, column=c).border = border

            fields = [
                ("HEX", row["hex"] or "N/A"),
                ("RGB", "N/A" if row["rgb"] is None else f"{row['r']}, {row['g']}, {row['b']}"),
                ("来源", row["source_quality"] or "N/A"),
            ]
            for offset, (label, value) in enumerate(fields, start=2):
                r = block_row + offset
                ws.merge_cells(
                    start_row=r,
                    start_column=block_col,
                    end_row=r,
                    end_column=block_col + tile_cols - 1,
                )
                cell = ws.cell(row=r, column=block_col, value=f"{label}: {value}")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
                for c in range(block_col, block_col + tile_cols):
                    ws.cell(row=r, column=c).border = border

    xlsx_path = out_dir / "colors.xlsx"
    wb.save(xlsx_path)
    normalize_xlsx_file(xlsx_path)


def draw_wrapped(draw: ImageDraw.ImageDraw, text: str, xy: tuple[int, int], font: ImageFont.ImageFont, fill: str, width: int, line_gap: int = 6) -> int:
    x, y = xy
    if not text:
        return y
    lines: list[str] = []
    for paragraph in text.splitlines():
        current = ""
        for char in paragraph:
            trial = current + char
            if draw.textbbox((0, 0), trial, font=font)[2] <= width or not current:
                current = trial
            else:
                lines.append(current)
                current = char
        lines.append(current)
    for line in lines:
        draw.text((x, y), line, font=font, fill=fill)
        y += font.size + line_gap
    return y


def checkerboard(draw: ImageDraw.ImageDraw, xy: tuple[int, int, int, int], size: int = 12) -> None:
    x1, y1, x2, y2 = xy
    colors = ("#F0F0F0", "#CFCFCF")
    for y in range(y1, y2, size):
        for x in range(x1, x2, size):
            draw.rectangle(
                (x, y, min(x + size, x2), min(y + size, y2)),
                fill=colors[((x - x1) // size + (y - y1) // size) % 2],
            )


def make_cover_page(info: dict[str, str], rows: list[dict[str, Any]], groups: OrderedDict[str, list[dict[str, Any]]]) -> Image.Image:
    market = info.get("market", {})
    page = Image.new("RGB", (1754, 1240), "#FFFFFF")
    draw = ImageDraw.Draw(page)
    draw.text((70, 70), info["title"], font=FONT_TITLE, fill="#111111")
    draw.text((70, 122), f"色号 + RGB 对照图例｜生成日期 {GENERATED_DATE}", font=FONT_SUBTITLE, fill="#555555")
    y = 190
    stats = [
        f"系列短名：{info['title']}",
        f"总颜色数：{len(rows)}",
        f"有 RGB：{sum(1 for row in rows if row['rgb'] is not None)}",
        f"无 RGB / 透明或未公开：{sum(1 for row in rows if row['rgb'] is None)}",
        f"分组：{', '.join(groups.keys())}",
        f"国内手工小店主流度：{market.get('tier', 'N/A')} / {market.get('score', 'N/A')} - {market.get('label', '')}",
    ]
    for line in stats:
        draw.text((90, y), line, font=FONT_BODY, fill="#222222")
        y += 42

    if market.get("summary"):
        y += 10
        draw.text((90, y), "主流度评估", font=FONT_SUBTITLE, fill="#111111")
        y += 40
        y = draw_wrapped(draw, market["summary"], (110, y), FONT_BODY, "#333333", 1450, 6)
        y += 10

    y += 20
    draw.text((90, y), "来源", font=FONT_SUBTITLE, fill="#111111")
    y += 42
    for url in source_urls(rows):
        y = draw_wrapped(draw, f"- {url}", (110, y), FONT_SMALL, "#333333", 1450, 4)
        y += 4

    y += 24
    draw.text((90, y), "说明", font=FONT_SUBTITLE, fill="#111111")
    y += 42
    notes = [
        "RGB 是屏幕参考值，不等于实物颜色；严谨对色仍建议以实物色卡为准。",
        "透明色、珠光、夜光等特殊材质在屏幕上不能完全表达，表内会保留色号并标注 RGB 是否缺失。",
        "公开源码库来源不等同于品牌官方标准；官方来源和第三方来源已在 README 与 JSON 中分开标注。",
    ]
    for note in notes:
        y = draw_wrapped(draw, f"- {note}", (110, y), FONT_BODY, "#333333", 1450, 6)
        y += 6
    return page


def make_color_pages(info: dict[str, str], groups: OrderedDict[str, list[dict[str, Any]]]) -> list[Image.Image]:
    pages: list[Image.Image] = []
    page_w, page_h = 1754, 1240
    margin_x, margin_y = 56, 58
    header_h = 90
    footer_h = 42
    tile_w, tile_h = 220, 112
    gap_x, gap_y = 15, 18
    cols = 7
    rows_per_page = 8
    per_page = cols * rows_per_page

    for group_name, group_rows_ in groups.items():
        total_pages = math.ceil(len(group_rows_) / per_page)
        for page_index in range(total_pages):
            page = Image.new("RGB", (page_w, page_h), "#FFFFFF")
            draw = ImageDraw.Draw(page)
            title = f"{info['title']} - {group_name}"
            draw.text((margin_x, 36), title, font=FONT_SUBTITLE, fill="#111111")
            draw.text(
                (margin_x, 70),
                f"{len(group_rows_)} 色｜第 {page_index + 1}/{total_pages} 页",
                font=FONT_SMALL,
                fill="#666666",
            )
            start = page_index * per_page
            for item_index, row in enumerate(group_rows_[start : start + per_page]):
                col = item_index % cols
                line = item_index // cols
                x = margin_x + col * (tile_w + gap_x)
                y = margin_y + header_h + line * (tile_h + gap_y)
                rgb = rgb_tuple(row)
                draw.rounded_rectangle((x, y, x + tile_w, y + tile_h), radius=8, outline="#B7B7B7", width=2, fill="#FFFFFF")
                swatch = (x + 8, y + 8, x + tile_w - 8, y + 50)
                if rgb is None:
                    checkerboard(draw, swatch)
                else:
                    draw.rectangle(swatch, fill=row["hex"])
                draw.rectangle(swatch, outline="#777777", width=1)
                code_fill = "#111111" if rgb is None else f"#{text_color_for(rgb)}"
                draw.text((x + 14, y + 16), row["code"], font=FONT_TILE_CODE, fill=code_fill)
                draw.text((x + 12, y + 58), row["hex"] or "HEX: N/A", font=FONT_TILE_SMALL, fill="#111111")
                draw.text((x + 12, y + 78), rgb_text(row), font=FONT_TILE_SMALL, fill="#111111")
                if row["transparency"]:
                    draw.text((x + 12, y + 96), "透明/特殊材质", font=FONT_TINY, fill="#666666")

            draw.line((margin_x, page_h - footer_h, page_w - margin_x, page_h - footer_h), fill="#DDDDDD", width=1)
            draw.text(
                (margin_x, page_h - footer_h + 12),
                "RGB 为参考显示值；特殊材质与实物颜色以实物色卡为准。",
                font=FONT_TINY,
                fill="#666666",
            )
            pages.append(page)
    return pages


def make_pdf(out_dir: Path, info: dict[str, str], rows: list[dict[str, Any]], groups: OrderedDict[str, list[dict[str, Any]]]) -> None:
    pages = [make_cover_page(info, rows, groups)]
    pages.extend(make_color_pages(info, groups))
    pdf_path = out_dir / "legend.pdf"
    first, rest = pages[0], pages[1:]
    first.save(
        pdf_path,
        "PDF",
        resolution=150.0,
        save_all=True,
        append_images=rest,
        title=info["title"],
        author="HansBug",
        creator="pindou-color-data",
        producer="Pillow",
        creationDate=FIXED_PDF_TIME,
        modDate=FIXED_PDF_TIME,
    )


def make_readme(out_dir: Path, info: dict[str, str], rows: list[dict[str, Any]], groups: OrderedDict[str, list[dict[str, Any]]]) -> None:
    market = info.get("market", {})
    quality_counts = Counter(row["source_quality"] or "unknown" for row in rows)
    missing = [row["code"] for row in rows if row["rgb"] is None]
    duplicate_codes = [code for code, count in Counter(row["code"] for row in rows).items() if count > 1]

    lines = [
        f"# {info['title']}",
        "",
        f"> 完整标题：{info.get('long_title', info['title'])}",
        "",
        f"- 生成日期：{GENERATED_DATE}",
        f"- 系列短名：{info['title']}",
        f"- 总颜色数：{len(rows)}",
        f"- 有 RGB：{sum(1 for row in rows if row['rgb'] is not None)}",
        f"- 无 RGB / 透明或未公开：{len(missing)}",
        f"- 分组：{', '.join(f'{name}({len(group)})' for name, group in groups.items())}",
        f"- 国内手工小店主流度：{market.get('tier', 'N/A')} / {market.get('score', 'N/A')} - {market.get('label', '')}",
        "",
        "## 文件",
        "",
        "- `colors.json`：脚本读取用，采用 `pindou-color-palette`；每个颜色含 `code`、`hex`、`rgb`、`group`、`source`。",
        "- `colors.xlsx`：人工查看用，不同色系分 sheet，色块 cell 已填充对应颜色。",
        "- `legend.pdf`：可直接转发的人类友好图例，包含色号、HEX、RGB 和实际色块。",
        "- `README.md`：本说明。",
        "",
        "## 国内手工小店主流度评估",
        "",
        f"- 评级：{market.get('tier', 'N/A')}",
        f"- 分数：{market.get('score', 'N/A')} / 5",
        f"- 标签：{market.get('label', '')}",
        f"- 摘要：{market.get('summary', '')}",
        "",
        "证据：",
        "",
    ]
    for evidence in market.get("evidence", []):
        lines.append(f"- {evidence}")

    lines.extend([
        "",
        "## 来源",
        "",
    ])
    for url in source_urls(rows):
        lines.append(f"- {url}")

    lines.extend(["", "## 来源质量统计", ""])
    for quality, count in quality_counts.items():
        lines.append(f"- `{quality}`：{count}")

    lines.extend(["", "## 注意事项", ""])
    lines.append("- RGB 是屏幕参考值，不等于实物颜色；严谨对色请以实物色卡为准。")
    if any(row["source_quality"].startswith("third_party") for row in rows):
        lines.append("- 本系列包含公开源码/工具站数据，不等同于品牌官方标准。")
    if "MARD" in info["title"] and "221" in info["title"]:
        lines.append("- MARD 公开来源之间存在 HEX 差异；顶层目录保留了 `Mard-221-source-differences.json` 供核对。")
    if missing:
        wrapped = textwrap.fill(", ".join(missing), width=100)
        lines.append(f"- 下列色号没有可用 RGB 或属于透明/特殊材质未公开 RGB：{wrapped}")
    if duplicate_codes:
        lines.append(f"- 检测到重复色号：{', '.join(duplicate_codes)}。JSON 中按原始顺序保留所有记录。")

    lines.extend(["", "## JSON 结构简例", "", "```json"])
    example = {
        "schema": SCHEMA,
        "id": info["id"],
        "title": info["title"],
        "count": len(rows),
        "colors": [
            {
                "code": rows[0]["code"] if rows else "",
                "hex": rows[0]["hex"] if rows else None,
                "rgb": None if not rows or rows[0]["rgb"] is None else [rows[0]["r"], rows[0]["g"], rows[0]["b"]],
                "group": next(iter(groups.keys()), ""),
            }
        ],
    }
    lines.append(json.dumps(example, ensure_ascii=False, indent=2))
    lines.extend(["```", ""])

    (out_dir / "README.md").write_text("\n".join(lines), encoding="utf-8")



def read_palette(directory: Path) -> tuple[dict[str, Any], list[dict[str, Any]], OrderedDict[str, list[dict[str, Any]]]]:
    data = json.loads((directory / "colors.json").read_text(encoding="utf-8"))
    if data.get("schema") != SCHEMA:
        raise ValueError(f"{directory} colors.json is not {SCHEMA}")
    info = {
        "id": data["id"],
        "title": data["title"],
        "long_title": data.get("description", data["title"]),
        "market": data.get("market", {}),
    }
    rows = normalize_rows(data.get("colors", []), data.get("sources", []))
    groups = group_rows(rows)
    for group_name, group in groups.items():
        for row in group:
            row["group"] = group_name
    return info, rows, groups


def make_index(manifest: list[dict[str, Any]]) -> None:
    lines = [
        "# 拼豆色卡四件套交付目录",
        "",
        f"生成日期：{GENERATED_DATE}",
        "",
        "每个子目录包含：`colors.json`、`colors.xlsx`、`legend.pdf`、`README.md`。",
        "",
        "本仓库定位为拼豆色卡数据仓库，适合其他应用以 git submodule 方式挂载使用。维护脚本保留在 `scripts/`，后续调整 JSON 后可复用脚本重新生成 XLSX/PDF/README。",
        "",
        "## 系列清单",
        "",
        "| 系列 | 主流度 | 说明 | 子目录 | 颜色数 | 有 RGB | 无 RGB |",
        "| --- | --- | --- | --- | ---: | ---: | ---: |",
    ]
    for item in sorted(manifest, key=lambda x: (-float(x.get("mainstream_score", 0)), x["id"])):
        lines.append(
            f"| [{item['title']}]({item['id']}/) | {item['mainstream_tier']} / {item['mainstream_score']} | {item['mainstream_label']} | `{item['id']}` | {item['count']} | {item['rows_with_rgb']} | {item['rows_without_rgb']} |"
        )
    lines.extend(
        [
            "",
            "## 主流度评级口径",
            "",
            "- S：国内手工小店/玩家图纸交流中的默认或最主流体系。",
            "- A：成熟、常见、有较强官方或工具生态支撑，但不一定是普通小店默认。",
            "- B：常见于工具、材料包或玩家圈，适合作为补充品牌。",
            "- 分数为 1-5 的相对评估，综合媒体报道、工具站覆盖、公开色卡生态、搜索可见度与官方资料。",
            "",
            "## 调研来源",
            "",
        ]
    )
    for src in MARKET_RESEARCH_SOURCES:
        lines.append(f"- {src['name']}：{src['url']}；{src['signal']}")

    lines.extend(
        [
            "",
            "附加文件：",
            "",
            "- `manifest.json`：所有子目录和统计的机器可读清单。",
            "- `Mard-221-source-differences.json`：MARD 221 两个公开来源之间的 HEX 差异。",
            "",
            "## 作为 Submodule 使用",
            "",
            "```bash",
            "git submodule add https://github.com/HansBug/pindou-color-data.git data/pindou-color-data",
            "```",
            "",
            "## 维护脚本",
            "",
            "- `scripts/generate_deliverables.py`：从 `colors.json` 重生成各系列四件套。",
            "- `scripts/build_tables.js`：上游采集/清洗辅助脚本。",
            "- 依赖：`python -m pip install openpyxl==3.1.5 pillow==12.2.0`。",
        ]
    )
    (ROOT / "README.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    (ROOT / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")



def main() -> None:
    manifest = []
    skip_dirs = {".git", "scripts", "__pycache__"}
    palette_dirs = sorted(
        p for p in ROOT.iterdir()
        if p.is_dir() and p.name not in skip_dirs and (p / "colors.json").exists()
    )
    for out_dir in palette_dirs:
        info, rows, groups = read_palette(out_dir)
        make_json(out_dir, info, rows, groups)
        make_xlsx(out_dir, info, rows, groups)
        make_pdf(out_dir, info, rows, groups)
        make_readme(out_dir, info, rows, groups)
        market = info.get("market", {})
        manifest.append({
            "id": info["id"],
            "title": info["title"],
            "description": info.get("long_title", info["title"]),
            "path": out_dir.name,
            "count": len(rows),
            "rows_with_rgb": sum(1 for row in rows if row["rgb"] is not None),
            "rows_without_rgb": sum(1 for row in rows if row["rgb"] is None),
            "mainstream_tier": market.get("tier", "N/A"),
            "mainstream_score": market.get("score", 0),
            "mainstream_label": market.get("label", ""),
            "mainstream_summary": market.get("summary", ""),
            "groups": {name: len(group) for name, group in groups.items()},
        })
    make_index(manifest)


if __name__ == "__main__":
    main()
